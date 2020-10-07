from openpyxl import Workbook
import requests
import json
from openpyxl.styles import Font, colors, Alignment
import demjson


# 向sheet中写入一行数据
def insertOne(value, sheet):
    row = [value] * 3
    sheet.append(row)


def getmidstring(html, start_str, end):
    start = html.find(start_str)
    if start >= 0:
        start += len(start_str)
        end = html.find(end, start)
        if end >= 0:
            return html[start:end].strip()

# 新建excel，并创建多个sheet
if __name__ == "__main__":


    f=input("请输入搜索内容")
    #创建工作簿抽象成活动，也就是我正在做这个时，记下来状态给 book
    book = Workbook()
    #获取json数组
    headers = {'User-Agent' : 'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; WOW64; Trident/4.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; .NET4.0C; InfoPath.3)'}
    res = requests.get("http://api.bilibili.com/x/web-interface/search/all/v2?keyword="+f,headers=headers)
    #取网页内容
    string = res.text
    #初始化循环初始值
    i = 0
    count = 1
    #sheet = book.create_sheet("sheet", 0) 意思是 创建表sheet并且插入到最前 把它抽象->活动对象 赋值给sheet
    #这里就不创建新表了  book.active 获取当前已有的工资表 抽象为->获得对象
    sheet = book.active
    #sheet['A1'] 是在访问 sheet 的 A1单元格，为空则写入文字 下面也是一样
    sheet['A1'] = '标题'
    #sheet.column_dimensions['A'] 给A列所有单元格 设置宽度 为 80 下面一样
    sheet.column_dimensions['A'].width=80
    sheet['B1'] = 'UID'
    sheet.column_dimensions['B'].width=20
    sheet['C1'] = 'BVID'
    sheet.column_dimensions['C'].width=20
    sheet['D1'] = '视频封面'
    sheet.column_dimensions['D'].width=80

    #还记得 string 吗？ 它是一个json数组 下面 demjson.decode(string) 是把他转换为 json对象
    list1 = demjson.decode(string)
    #确定对象
    listarr = list1['data']['result'][8]['data']
    #现在我们开始循环，把对象里的值抽出来 len(listarr)是获取数组长度
    while i<len(listarr):
        #下面我们进入循环 当 i的数值比数组长度小时，就执行下面，否则跳出循环 下面的i是在定位每一组数组对象
        arr = list1['data']['result'][8]['data'][i]
        count = count + 1
        i = i+1
        # sheet.cell(count, 1, arr['title']) 传入的3个值 行数 列数 值 如果单元格为空则创建写入这些内容
        sheet.cell(count, 1, arr['title'])
        #这里.alignment = Alignment(horizontal='center', vertical='center') 是给这个单元格设置格式为居中
        sheet.cell(i, 1).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(count, 2, arr['mid'])
        sheet.cell(i, 2).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(count, 3, arr['bvid'])
        sheet.cell(i, 3).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(count, 4, "http:"+arr['pic'])
        sheet.cell(i, 4).alignment = Alignment(horizontal='center', vertical='center')



        

    

        
    # 每个sheet里设置列标题
    '''
    for i in range(0, 2):
        # 为每个sheet设置title，插入位置index
        sheet = book.create_sheet("sheet" + str(i + 1), i)
        # 每个sheet里设置列标题
        sheet.append(["title" + str(i + 1)] * 3)
    '''


    # 向sheet中插入数据
    '''
    for i in range(0, 10):
        insertOne("ni", book.get_sheet_by_name(sheets[0]))
        insertOne("wo", book.get_sheet_by_name(sheets[0]))
        insertOne("ta", book.get_sheet_by_name(sheets[0]))
        count = count + 1
    '''


    # 保存数据到.xlsx文件
    book.save("test.xlsx")
    print(str(count))