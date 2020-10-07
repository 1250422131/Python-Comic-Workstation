import tkinter
import easygui as g
from tkinter import *
#导入表格模块
from openpyxl import Workbook
from openpyxl.styles import Font, colors, Alignment
#网络模块导入
import requests
#json模块导入
import demjson
#正则模块导入
import re
import time






if __name__ == "__main__":
    BV = g.enterbox(msg="请输入番剧BV号，已,分割BV",title="弹幕下载")
    print(BV)
    #新建一个excel活动
    book = Workbook()
    #爬虫头设置
    headers = {'User-Agent' : 'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; WOW64; Trident/4.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; .NET4.0C; InfoPath.3)'}
    #bv_i 分P遍历记录值
    bv_i = 0
    #弹幕遍历记录值
    i = 0
    #弹幕对应行数遍历记录
    i_n = 1
    if BV == None:
        sys.exit(0)
    elif BV == "":
        g.msgbox(msg="没有输入内容！",title="提示",ok_button="确定")
        sys.exit(0)
    #待抓取视频数组
    bv_sz = BV.split(',')
    #BV遍历记录
    bv_sz_i = 0
    #循环跑每一个BV
    while bv_sz_i < len(bv_sz):
        time.sleep(4)
        #新建一个活动表 名称为BV号
        sheet = book.create_sheet(bv_sz[bv_sz_i], 0)
        #cid数据获取
        BV_res = requests.get("https://api.bilibili.com/x/player/pagelist?bvid="+str(bv_sz[bv_sz_i])+"&jsonp=jsonp",headers=headers)
        BV_res.encoding = 'utf-8'
        BV_string = BV_res.text
        #print(bv_sz[bv_sz_i])
        bv_sz_i = bv_sz_i + 1
        BV_Strjson = demjson.decode(BV_string)
        BV_CID = BV_Strjson['data']
        #循环获取每一个分P视频
        print("当前第"+str(bv_sz_i)+"个,共"+str(len(bv_sz)))
        while bv_i < len(BV_CID):
            #cid入表
            # sheet.cell(count, 1, arr['title']) 传入的3个值 行数 列数 值 如果单元格为空则创建写入这些内容
            # 即代表第一行横着依次写入分P的CID
            sheet.cell(1, bv_i+1, BV_CID[bv_i]['cid'])
            # 单元格水平垂直居中
            sheet.cell(1, bv_i+1).alignment = Alignment(horizontal='center', vertical='center')
            #调用输出 print(BV_CID[bv_i]['cid'])
            #获取弹幕信息
            res = requests.get("http://api.bilibili.com/x/v1/dm/list.so?oid="+str(BV_CID[bv_i]['cid']),headers=headers)
            res.encoding = 'utf-8'
            bv_i = bv_i + 1
            #取弹幕信息内容
            string = res.text
            #正则匹配->全体数据匹配
            DMData = re.compile('<d.*?>(.*?)</d>')
            #执行正则表达式->贪婪获取数据
            DMArray = DMData.findall(string)
            while i < len(DMArray):
                #不要覆盖CID 即代表从第二行录入 列为分P对应的数列值
                sheet.cell(i_n+1, bv_i, DMArray[i])
                #设置水平垂直居中
                sheet.cell(i_n+1, bv_i).alignment = Alignment(horizontal='center', vertical='center')
                #print(DMArray[i])
                i= i + 1
                #数据下滚->避免把上一个数据覆盖了
                i_n = i + 1
            i = 0
        bv_i = 0
    book.save("dm1.xlsx")
print(i)
